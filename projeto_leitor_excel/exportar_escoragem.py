#!/usr/bin/env python3
"""
exportar_escoragem.py
Gera Excel e preview HTML a partir de JSON de motor de escoragem.
"""

import html as _html
import io
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_XL_RED    = "FF4444"
_XL_YELLOW = "FFFF00"
_XL_LGREEN = "92D050"
_XL_GREEN  = "00B050"

_CSS_COLORS = {
    _XL_RED:    "#FF4444",
    _XL_YELLOW: "#FFFF00",
    _XL_LGREEN: "#92D050",
    _XL_GREEN:  "#00B050",
}

# Fundos claros exigem texto PRETO; fundos escuros exigem texto BRANCO
_LIGHT_BG_XL  = {_XL_YELLOW, _XL_LGREEN, "FFC000", "FFFFFF", "EBF3FA", "D6E4F0"}
_LIGHT_BG_CSS = {"#FFFF00", "#92D050", "#FFC000", "#FFFFFF"}

_RISK_COLORS_XL = {
    "muito baixo": _XL_GREEN,
    "baixo":       _XL_LGREEN,
    "médio":       _XL_YELLOW,
    "medio":       _XL_YELLOW,
    "alto":        "FFC000",
    "muito alto":  _XL_RED,
}
_RISK_COLORS_CSS = {
    "muito baixo": "#00B050",
    "baixo":       "#92D050",
    "médio":       "#FFFF00",
    "medio":       "#FFFF00",
    "alto":        "#FFC000",
    "muito alto":  "#FF4444",
}

_THIN   = Side(style="thin",   color="AAAAAA")
_MEDIUM = Side(style="medium", color="1F4E79")
_BTHIN  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _font_color_xl(bg: str) -> str:
    """Retorna cor de fonte Excel para o fundo dado."""
    return "000000" if bg in _LIGHT_BG_XL else "FFFFFF"


def _font_color_css(bg_css: str) -> str:
    """Retorna cor de fonte CSS para o fundo dado."""
    return "#000000" if bg_css in _LIGHT_BG_CSS else "#FFFFFF"


def _score_color_xl(score, max_score) -> str:
    try:
        s, m = float(score), float(max_score)
    except (TypeError, ValueError):
        return "FFFFFF"
    if m <= 0:
        return "FFFFFF"
    if s == 0:
        return _XL_RED
    pct = s / m
    if pct < 0.6:
        return _XL_YELLOW
    if pct < 1.0:
        return _XL_LGREEN
    return _XL_GREEN


_HTML_ENT = {"&gt;=": ">=", "&lt;=": "<=", "&gt;": ">", "&lt;": "<", "&amp;": "&"}


def _extract_rich_text(obj) -> str:
    """Extrai texto plano de objeto ProseMirror/rich-text doc."""
    if isinstance(obj, str):
        return obj
    if isinstance(obj, dict):
        if obj.get("type") == "text":
            return obj.get("text", "")
        return " ".join(
            _extract_rich_text(child) for child in obj.get("content", [])
        ).strip()
    if isinstance(obj, list):
        return " ".join(_extract_rich_text(item) for item in obj).strip()
    return str(obj) if obj is not None else ""


def _decode(text) -> str:
    if text is None:
        return ""
    if isinstance(text, dict):
        return _extract_rich_text(text)
    if not isinstance(text, str):
        return str(text)
    for k, v in _HTML_ENT.items():
        text = text.replace(k, v)
    return _html.unescape(text)


def _fmt_num(v):
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _to_str(v) -> str:
    if v is None or v == "":
        return ""
    return str(_fmt_num(v))


def build_condition(cond1=None, val1=None, cond2=None, val2=None) -> str:
    c1 = _decode(cond1)
    c2 = _decode(cond2)
    if isinstance(val1, (dict, list)):
        val1 = _extract_rich_text(val1)
    if isinstance(val2, (dict, list)):
        val2 = _extract_rich_text(val2)
    v1 = _fmt_num(val1)
    v2 = _fmt_num(val2)
    if not c1:
        return ""
    if c1.lower() == "vazio":
        return "vazio"
    if c1 == "==" and (v2 == "" or v2 is None):
        return f"= {v1}"
    if not c2 or v2 == "" or v2 is None:
        return f"{c1} {v1}"
    return f"{c1} {v1} e {c2} {v2}"


def _fmt_peso(peso) -> str:
    if isinstance(peso, float) and 0 < peso <= 1:
        return f"{peso * 100:.0f}%"
    return f"{peso}%" if peso not in ("", None, 0) else ""


def _fmt_validade(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)) and val > 0:
        return f"{int(val)} dias"
    return str(val)


def _peso_to_float(peso) -> float:
    """Normaliza peso para decimal (0.0–1.0), independente de como foi armazenado."""
    if isinstance(peso, float) and 0 < peso <= 1:
        return peso
    try:
        v = float(peso)
        return v / 100.0 if v >= 1 else v
    except (TypeError, ValueError):
        return 0.0


def safe_filename(text: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text).strip() or "escoragem"


def _clean_var_name(name: str) -> str:
    """Remove sufixos de I.A do nome da variavel para exibicao."""
    name = re.sub(r'\s*[-–]\s*I\.?\s*A\.?\s*$', '', name, flags=re.IGNORECASE)
    return name.strip()


def _parse_grupos(data: dict) -> List[Dict]:
    result = []
    for grupo in data.get("grupos", []):
        variaveis = []
        for var in grupo.get("variaveis", []):
            pts = var.get("pontuacao", [])
            nums = []
            for p in pts:
                s = p.get("faixa_pontuacao")
                if s is not None and s != "":
                    try:
                        nums.append(float(s))
                    except (ValueError, TypeError):
                        pass
            max_s = max(nums) if nums else 0
            variaveis.append({
                "nome":       _clean_var_name(var.get("nome", "")),
                "aplicar_a":  var.get("aplicar_a", ""),
                "peso":       var.get("peso", 0),
                "max_score":  max_s,
                "pontuacoes": [
                    {
                        "condicao": build_condition(
                            p.get("condicao_1"), p.get("pontuacao_1"),
                            p.get("condicao_2"), p.get("pontuacao_2"),
                        ),
                        "score": _fmt_num(p.get("faixa_pontuacao", 0)),
                    }
                    for p in pts
                ],
            })
        result.append({"nome": grupo.get("nome", ""), "variaveis": variaveis})
    return result


def _parse_classificacao(data: dict) -> List[Dict]:
    result = []
    for cls in data.get("classificacao", []):
        classe   = cls.get("classe") or cls.get("nome", "")
        risco    = cls.get("risco", "")
        condicao = build_condition(
            cls.get("condicao_1"), cls.get("pontuacao_1"),
            cls.get("condicao_2"), cls.get("pontuacao_2"),
        )
        fator = _fmt_num(cls.get("faixa_pontuacao", ""))
        result.append({
            "classe":          classe,
            "risco":           risco,
            "condicao":        condicao,
            "fator":           _to_str(fator),
            "validade":        _fmt_validade(cls.get("validade_limite")),
            "limite_sugerido": _to_str(cls.get("limite_sugerido", "")),
        })
    return result


_COND_OPS = r">=|<=|==|!=|>|<"


def _parse_value_token(tok: str):
    tok = tok.strip()
    try:
        v = float(tok)
        return int(v) if v == int(v) else v
    except ValueError:
        return tok


def _parse_condition_text(text) -> Tuple[Optional[str], object, Optional[str], object]:
    text = (text or "").strip()
    if not text:
        return None, None, None, None
    if text.lower() == "vazio":
        return "vazio", None, None, None
    if text.startswith("= "):
        return "==", _parse_value_token(text[2:]), None, None
    m = re.match(rf"^({_COND_OPS})\s*(\S+)\s+e\s+({_COND_OPS})\s*(\S+)$", text)
    if m:
        c1, v1, c2, v2 = m.groups()
        return c1, _parse_value_token(v1), c2, _parse_value_token(v2)
    m = re.match(rf"^({_COND_OPS})\s*(\S+)$", text)
    if m:
        c1, v1 = m.groups()
        return c1, _parse_value_token(v1), None, None
    return None, text, None, None


def _parse_peso_pct(text) -> float:
    if text is None:
        return 0.0
    s = str(text).strip().replace(",", ".").replace("%", "")
    try:
        return float(s) / 100.0
    except ValueError:
        return 0.0


def _parse_validade_text(text) -> Optional[int]:
    m = re.match(r"(\d+)", str(text or "").strip())
    return int(m.group(1)) if m else None


def _ffill(values: list) -> list:
    out, last = [], None
    for v in values:
        if v not in (None, ""):
            last = v
        out.append(last)
    return out


def _read_grupos_sheet(ws) -> List[Dict]:
    rows = []
    for r in range(4, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        if isinstance(col_a, str) and col_a.strip().upper() == "TOTAL":
            break
        rows.append({
            "grupo":  col_a,
            "var":    ws.cell(r, 2).value,
            "faixa":  ws.cell(r, 3).value,
            "pontos": ws.cell(r, 4).value,
            "max":    ws.cell(r, 5).value,
            "peso":   ws.cell(r, 6).value,
        })

    grupo_col = _ffill([row["grupo"] for row in rows])
    var_col   = _ffill([row["var"]   for row in rows])
    max_col   = _ffill([row["max"]   for row in rows])
    peso_col  = _ffill([row["peso"]  for row in rows])

    grupos: Dict[str, Dict] = {}
    for i, row in enumerate(rows):
        nome_grupo = grupo_col[i] or ""
        nome_var   = var_col[i] or ""
        grupo = grupos.setdefault(nome_grupo, {"nome": nome_grupo, "variaveis": {}})
        var = grupo["variaveis"].setdefault(nome_var, {
            "nome": nome_var,
            "peso": _parse_peso_pct(peso_col[i]),
            "pontuacao": [],
        })

        try:
            pontos_max_v = float(max_col[i]) if max_col[i] not in (None, "") else 0.0
        except (TypeError, ValueError):
            pontos_max_v = 0.0
        try:
            pf = float(row["pontos"]) if row["pontos"] not in (None, "") else 0.0
        except (TypeError, ValueError):
            pf = 0.0
        pct = round(pf / pontos_max_v * 100, 4) if pontos_max_v else 0.0
        if pct == int(pct):
            pct = int(pct)

        c1, v1, c2, v2 = _parse_condition_text(row["faixa"])
        var["pontuacao"].append({
            "condicao_1": c1,
            "pontuacao_1": v1,
            "condicao_2": c2,
            "pontuacao_2": v2,
            "faixa_pontuacao": pct,
        })

    return [
        {"nome": g["nome"], "variaveis": list(g["variaveis"].values())}
        for g in grupos.values()
    ]


def _read_classificacao_sheet(ws) -> List[Dict]:
    cls_rows = []
    for r in range(3, ws.max_row + 1):
        classe = ws.cell(r, 1).value
        if classe in (None, ""):
            break
        cls_rows.append({
            "classe":   classe,
            "risco":    ws.cell(r, 2).value,
            "faixa":    ws.cell(r, 3).value,
            "validade": ws.cell(r, 4).value,
        })

    sep_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "RESUMO" in v.upper():
            sep_row = r
            break

    extras: Dict[object, Dict] = {}
    if sep_row:
        for r in range(sep_row + 2, ws.max_row + 1):
            classe = ws.cell(r, 1).value
            if classe in (None, ""):
                break
            extras[classe] = {
                "fator":  ws.cell(r, 3).value,
                "limite": ws.cell(r, 4).value,
            }

    result = []
    for row in cls_rows:
        c1, v1, c2, v2 = _parse_condition_text(row["faixa"])
        extra = extras.get(row["classe"], {})
        fator = extra.get("fator")
        limite_txt = extra.get("limite")
        if isinstance(limite_txt, str) and limite_txt.strip().lower() == "fator x capacidade de pagamento":
            limite_txt = ""
        result.append({
            "classe":          row["classe"],
            "risco":           row["risco"],
            "condicao_1":      c1,
            "pontuacao_1":     v1,
            "condicao_2":      c2,
            "pontuacao_2":     v2,
            "faixa_pontuacao": _parse_value_token(str(fator)) if fator not in (None, "", "—") else "",
            "validade_limite": _parse_validade_text(row["validade"]),
            "limite_sugerido": limite_txt or "",
        })
    return result


def import_excel_to_json(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "EscoreSugerido" not in wb.sheetnames:
        raise ValueError("Aba 'EscoreSugerido' não encontrada na planilha.")

    ws_score   = wb["EscoreSugerido"]
    model_name = ws_score["A1"].value or ""
    grupos     = _read_grupos_sheet(ws_score)

    classificacao = []
    if "Classificacao de Risco" in wb.sheetnames:
        classificacao = _read_classificacao_sheet(wb["Classificacao de Risco"])

    return {"detalhes": model_name, "grupos": grupos, "classificacao": classificacao}


def _hdr(cell, text: str, bg: str = "1F4E79", fg: str = "FFFFFF",
         bold: bool = True, rot: int = 0) -> None:
    cell.value     = text
    cell.font      = Font(bold=bold, color=fg, name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, text_rotation=rot)
    cell.border    = _BTHIN


def _dat(cell, value, bg: str = "", bold: bool = False,
         align: str = "center", wrap: bool = False, font_color: str = "") -> None:
    if isinstance(value, str) and value.startswith("="):
        # Evita que openpyxl trate a string como fórmula Excel (#NAME? / #VALUE!)
        cell.value = value
        cell.data_type = "s"
    else:
        cell.value = value
    fc = font_color or (_font_color_xl(bg) if bg else "000000")
    cell.font      = Font(name="Calibri", size=10, bold=bold, color=fc)
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = _BTHIN


def _write_escoragem_sheet(ws, grupos: List[Dict], model_name: str) -> None:
    _N = 6
    _ALT_BG = "EBF3FA"   # azul muito claro para linhas alternadas
    _SEP_BG = "D6E4F0"   # azul claro para subtítulo

    soma_pesos   = sum(_peso_to_float(v["peso"]) for g in grupos for v in g["variaveis"])
    total_pontos = soma_pesos * 1000

    ws.merge_cells(f"A1:{get_column_letter(_N)}1")
    c = ws["A1"]
    c.value     = model_name
    c.font      = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = Border(bottom=Side(style="medium", color="AAAAAA"))
    ws.row_dimensions[1].height = 40

    ws.merge_cells(f"A2:{get_column_letter(_N)}2")
    s = ws["A2"]
    s.value = (
        f"Soma dos Pesos: {soma_pesos * 100:.0f}%"
        f"     |     Pontuação Máxima Total: {_fmt_num(total_pontos)} pts"
    )
    s.font      = Font(italic=True, size=9, name="Calibri", color="1F4E79")
    s.fill      = PatternFill("solid", fgColor=_SEP_BG)
    s.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    s.border    = Border(bottom=Side(style="thin", color="B0C8E0"))
    ws.row_dimensions[2].height = 18

    hdrs = [
        ("Grupo",               "1F4E79", "FFFFFF"),
        ("Variáveis\nEscore",   "1F4E79", "FFFFFF"),
        ("Faixa de\nPontuação", "1F4E79", "FFFFFF"),
        ("Pontos\nFaixa",       "C00000", "FFFFFF"),
        ("Pontos\nMáx",         "C00000", "FFFFFF"),
        ("Peso\n(0-100)%",      "C00000", "FFFFFF"),
    ]
    for col, (txt, bg, fg) in enumerate(hdrs, 1):
        _hdr(ws.cell(3, col), txt, bg=bg, fg=fg)
    ws.row_dimensions[3].height = 32

    cur = 4
    row_parity = 0

    for grupo in grupos:
        g_start = cur

        for var in grupo["variaveis"]:
            v_start      = cur
            max_s        = var["max_score"]
            pontos_max_v = _peso_to_float(var["peso"]) * 1000
            pontos_max_s = str(_fmt_num(pontos_max_v)) if pontos_max_v else ""

            for regra in var["pontuacoes"]:
                score = regra["score"]
                alt   = _ALT_BG if row_parity % 2 == 0 else ""
                try:
                    sf   = float(score) if score != "" else 0.0
                    xcol = _score_color_xl(sf, max_s)
                    pf   = (sf / max_s) * pontos_max_v if max_s > 0 else 0.0
                    pfs  = _fmt_num(pf)
                except (TypeError, ValueError):
                    xcol, pfs = "FFFFFF", ""

                _dat(ws.cell(cur, 3), regra["condicao"], align="left", bg=alt)
                _dat(ws.cell(cur, 4), pfs, bg=xcol, bold=True)
                row_parity += 1
                cur += 1

            v_end = cur - 1
            for col in (2, 5, 6):
                if v_end > v_start:
                    ws.merge_cells(
                        f"{get_column_letter(col)}{v_start}:{get_column_letter(col)}{v_end}"
                    )
            _dat(ws.cell(v_start, 2), var["nome"], align="center", wrap=True)
            _dat(ws.cell(v_start, 5), pontos_max_s, align="center",
                 bold=True, font_color="1F4E79")
            _dat(ws.cell(v_start, 6), _fmt_peso(var["peso"]), align="center")

        g_end = cur - 1
        if g_end > g_start:
            ws.merge_cells(f"A{g_start}:A{g_end}")
        gc = ws.cell(g_start, 1)
        gc.value     = grupo["nome"]
        gc.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        gc.fill      = PatternFill("solid", fgColor="2E75B6")
        gc.alignment = Alignment(horizontal="center", vertical="center",
                                  text_rotation=90, wrap_text=False)
        gc.border    = Border(left=_MEDIUM, right=_THIN, top=_MEDIUM, bottom=_MEDIUM)

    total_row = cur
    ws.merge_cells(f"A{total_row}:D{total_row}")
    tc = ws.cell(total_row, 1)
    tc.value     = "TOTAL"
    tc.font      = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _BTHIN
    _dat(ws.cell(total_row, 5), _fmt_num(total_pontos), bg="1F4E79",
         bold=True, font_color="FFFFFF")
    _dat(ws.cell(total_row, 6), f"{soma_pesos * 100:.0f}%", bg="1F4E79",
         font_color="FFFFFF")
    ws.row_dimensions[total_row].height = 20

    for col, w in enumerate([8, 32, 32, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C4"


def _write_classificacao_sheet(ws, classificacao: List[Dict]) -> None:
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "CLASSIFICACAO DE RISCO"
    c.font      = Font(bold=True, size=14, name="Calibri", color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for col, hdr in enumerate(["Classe", "Risco", "Faixa de Pontuacao", "Validade"], 1):
        _hdr(ws.cell(2, col), hdr)
    ws.row_dimensions[2].height = 24

    for ri, cls in enumerate(classificacao, 3):
        risco_key = cls["risco"].lower().replace("é", "e")
        bg = _RISK_COLORS_XL.get(risco_key, "FFFFFF")
        _dat(ws.cell(ri, 1), cls["classe"],   bg=bg, bold=True)
        _dat(ws.cell(ri, 2), cls["risco"],    bg=bg)
        _dat(ws.cell(ri, 3), cls["condicao"])
        _dat(ws.cell(ri, 4), cls["validade"])
        ws.row_dimensions[ri].height = 20

    last_cls_row = 2 + len(classificacao)

    sep_row = last_cls_row + 2

    ws.merge_cells(f"A{sep_row}:D{sep_row}")
    t = ws.cell(sep_row, 1)
    t.value     = "RESUMO — CALCULO DE LIMITE SUGERIDO"
    t.font      = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.border    = _BTHIN
    ws.row_dimensions[sep_row].height = 26

    sub_row = sep_row + 1
    for col, hdr in enumerate(["Classe", "Risco", "Fator de Limite", "Limite Sugerido"], 1):
        _hdr(ws.cell(sub_row, col), hdr, bg="2E75B6")
    ws.row_dimensions[sub_row].height = 22

    for i, cls in enumerate(classificacao):
        ri2 = sub_row + 1 + i
        risco_key = cls["risco"].lower().replace("é", "e")
        bg = _RISK_COLORS_XL.get(risco_key, "FFFFFF")
        _dat(ws.cell(ri2, 1), cls["classe"],          bg=bg, bold=True)
        _dat(ws.cell(ri2, 2), cls["risco"],           bg=bg)
        _dat(ws.cell(ri2, 3), cls["fator"],           align="center")
        limite_txt = cls["limite_sugerido"] if cls["limite_sugerido"] else "Fator x Capacidade de Pagamento"
        _dat(ws.cell(ri2, 4), limite_txt, align="left")
        ws.row_dimensions[ri2].height = 20

    nota_row = sub_row + 1 + len(classificacao) + 1
    ws.merge_cells(f"A{nota_row}:D{nota_row + 2}")
    nota = ws.cell(nota_row, 1)
    nota.value = (
        "Como interpretar:\n"
        "O Limite Sugerido e calculado aplicando o Fator de Limite da classe do cliente "
        "sobre a Capacidade de Pagamento Mensal apurada no motor.\n"
        "Exemplo: Classe A — Fator 1 — Limite = 1 x Capacidade de Pagamento."
    )
    nota.font      = Font(name="Calibri", size=9, italic=True, color="444444")
    nota.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
    nota.border    = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    nota.fill = PatternFill("solid", fgColor="F5F5F5")
    for r in range(nota_row, nota_row + 3):
        ws.row_dimensions[r].height = 18

    for col, w in enumerate([16, 16, 28, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_workbook(data: dict) -> Tuple[str, Workbook]:
    model_name    = data.get("detalhes", "escoragem")
    grupos        = _parse_grupos(data)
    classificacao = _parse_classificacao(data)

    wb = Workbook()
    wb.remove(wb.active)

    _write_escoragem_sheet(wb.create_sheet("EscoreSugerido"), grupos, model_name)
    _write_classificacao_sheet(wb.create_sheet("Classificacao de Risco"), classificacao)

    return safe_filename(model_name) + ".xlsx", wb


def export_to_bytes(json_bytes: bytes) -> Tuple[str, bytes]:
    data      = json.loads(json_bytes)
    fname, wb = _build_workbook(data)
    buf       = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return fname, buf.read()


def export_json_to_excel(json_path: Path, output_dir: Optional[Path] = None) -> Path:
    with json_path.open(encoding="utf-8") as f:
        data = json.load(f)
    fname, wb = _build_workbook(data)
    out_dir   = output_dir or json_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path  = out_dir / fname
    wb.save(out_path)
    print(f"Exportado: {out_path}")
    return out_path


_BASE_CSS = """
<style>
  /* Reset — garante fundo branco e texto preto independente do tema do Streamlit */
  html, body { background: #ffffff; color: #000000; }
  * { box-sizing: border-box; margin: 0; padding: 0; }

  .esc-wrap {
    overflow-x: auto;
    overflow-y: auto;
    max-height: 640px;
    border: 1px solid #D0D7E3;
    border-radius: 8px;
    background: #ffffff;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
  }
  .esc-tbl {
    border-collapse: collapse;
    font-family: Calibri, Arial, sans-serif;
    font-size: 12px;
    width: 100%;
    min-width: 700px;
    background: #ffffff;
  }
  /* Todas as celulas de dados: texto PRETO por padrao */
  .esc-tbl td {
    border: 1px solid #D0D7E3;
    padding: 5px 9px;
    text-align: center;
    vertical-align: middle;
    color: #000000;
    background: #ffffff;
  }
  /* Cabecalhos: texto BRANCO */
  .esc-tbl th {
    border: 1px solid #1A406A;
    padding: 6px 9px;
    text-align: center;
    vertical-align: middle;
    background: #1F4E79;
    color: #ffffff;
    font-weight: bold;
    white-space: nowrap;
  }
  .hdr-score { background: #C00000 !important; color: #ffffff !important; }

  .esc-title {
    background: #F0F4FA;
    color: #1F4E79;
    font-size: 13px;
    font-weight: bold;
    text-align: center;
    padding: 10px;
    letter-spacing: 0.5px;
    border-radius: 8px 8px 0 0;
  }
  .esc-grupo {
    background: #2E75B6;
    color: #ffffff;
    font-weight: bold;
    writing-mode: vertical-lr;
    transform: rotate(180deg);
    white-space: nowrap;
    width: 26px;
    padding: 8px 3px;
  }
  .esc-var {
    text-align: left;
    padding-left: 10px;
    min-width: 150px;
    max-width: 220px;
    white-space: normal;
    word-break: break-word;
    color: #000000;
    background: #ffffff;
  }
  .esc-faixa {
    text-align: left;
    padding-left: 8px;
    min-width: 130px;
    color: #000000;
    background: #ffffff;
  }
  /* Linhas alternadas leves */
  .esc-tbl tbody tr:nth-child(even) td:not([style*="background"]) {
    background: #F7F9FC;
  }
  /* Separa visualmente um grupo do próximo */
  .esc-tbl tbody tr.group-end td {
    border-bottom: 3px solid #1F4E79;
  }
</style>
"""


def build_preview_html(data: dict) -> str:
    grupos     = _parse_grupos(data)
    model_name = data.get("detalhes", "Escoragem")
    has_aplic  = any(v["aplicar_a"] for g in grupos for v in g["variaveis"])

    n_cols = 7 + (1 if has_aplic else 0)

    soma_pesos   = sum(_peso_to_float(v["peso"]) for g in grupos for v in g["variaveis"])
    total_pontos = soma_pesos * 1000
    soma_fmt     = f"{soma_pesos * 100:.0f}%"
    total_fmt    = str(_fmt_num(total_pontos)) if total_pontos else "—"

    th  = "<th>Grupo</th><th>Variaveis Escore</th>"
    th += "<th>Aplicar a</th>" if has_aplic else ""
    th += "<th>Faixa de Pontuacao</th>"
    th += '<th class="hdr-score">% do Max</th>'
    th += '<th class="hdr-score">Pontos</th>'
    th += '<th class="hdr-score">Pontos Max</th>'
    th += '<th class="hdr-score">Peso (0-100)%</th>'

    rows = []
    for grupo in grupos:
        g_total   = sum(len(v["pontuacoes"]) for v in grupo["variaveis"])
        g_done    = False
        g_emitted = 0

        for var in grupo["variaveis"]:
            v_cnt        = len(var["pontuacoes"])
            max_s        = var["max_score"]
            v_first      = True
            pontos_max_v = _peso_to_float(var["peso"]) * 1000
            pontos_max_s = str(_fmt_num(pontos_max_v)) if pontos_max_v else ""

            for regra in var["pontuacoes"]:
                score = regra["score"]
                try:
                    sf    = float(score) if score != "" else 0.0
                    xkey  = _score_color_xl(sf, max_s)
                    bg_c  = _CSS_COLORS.get(xkey, "#FFFFFF")
                    fc    = _font_color_css(bg_c)
                    pct   = sf / max_s * 100 if max_s > 0 else 0.0
                    ptxt  = f"{pct:.1f}%"
                    sdsp  = str(_fmt_num(score)) if score != "" else "0"
                    pf    = (sf / max_s) * pontos_max_v if max_s > 0 else 0.0
                    pftxt = str(_fmt_num(pf))
                except (TypeError, ValueError):
                    bg_c, fc, ptxt, sdsp, pftxt = "#FFFFFF", "#000000", "", str(score), ""

                g_emitted += 1
                row_class = ' class="group-end"' if g_emitted == g_total else ""
                r = f"<tr{row_class}>"
                if not g_done:
                    r += f'<td class="esc-grupo" rowspan="{g_total}">{_html.escape(grupo["nome"])}</td>'
                    g_done = True

                if v_first:
                    r += f'<td class="esc-var" rowspan="{v_cnt}">{_html.escape(var["nome"])}</td>'
                    if has_aplic:
                        r += f'<td rowspan="{v_cnt}">{_html.escape(var["aplicar_a"])}</td>'

                r += f'<td class="esc-faixa">{_html.escape(regra["condicao"])}</td>'
                r += f'<td style="background:{bg_c};color:{fc}">{ptxt}</td>'
                r += f'<td style="background:{bg_c};color:{fc};font-weight:bold">{pftxt}</td>'

                if v_first:
                    r += f'<td rowspan="{v_cnt}" style="font-weight:bold;color:#1F4E79">{pontos_max_s}</td>'
                    r += f'<td rowspan="{v_cnt}">{_fmt_peso(var["peso"])}</td>'
                    v_first = False

                r += "</tr>"
                rows.append(r)

    return f"""
{_BASE_CSS}
<style>
  .esc-subtitle {{
    background: #D6E4F0;
    color: #1F4E79;
    font-size: 11px;
    text-align: center;
    padding: 6px 10px;
    border-bottom: 1px solid #B0C8E0;
  }}
  .esc-subtitle strong {{ color: #C00000; }}
</style>
<div class="esc-wrap">
  <table class="esc-tbl">
    <thead>
      <tr><td class="esc-title" colspan="{n_cols}">
        {_html.escape(model_name)}
      </td></tr>
      <tr><td class="esc-subtitle" colspan="{n_cols}">
        Soma dos Pesos: <strong>{soma_fmt}</strong>
        &nbsp;&nbsp;|&nbsp;&nbsp;
        Pontuação Máxima Total: <strong>{total_fmt} pts</strong>
      </td></tr>
      <tr>{th}</tr>
    </thead>
    <tbody>{"".join(rows)}</tbody>
  </table>
</div>
"""


def build_classificacao_html(data: dict) -> str:
    classificacao = _parse_classificacao(data)
    model_name    = data.get("detalhes", "Escoragem")

    if not classificacao:
        return "<p>Sem dados de classificacao.</p>"

    cls_rows = []
    for cls in classificacao:
        risco_key = cls["risco"].lower().replace("é", "e")
        bg  = _RISK_COLORS_CSS.get(risco_key, "#FFFFFF")
        fc  = _font_color_css(bg)
        cls_rows.append(
            f"<tr>"
            f'<td style="background:{bg};color:{fc};font-weight:bold">{_html.escape(cls["classe"])}</td>'
            f'<td style="background:{bg};color:{fc}">{_html.escape(cls["risco"])}</td>'
            f'<td>{_html.escape(cls["condicao"])}</td>'
            f'<td>{_html.escape(cls["validade"])}</td>'
            f"</tr>"
        )

    has_limite = any(cls["fator"] or cls["limite_sugerido"] for cls in classificacao)
    lim_rows = []
    for cls in classificacao:
        risco_key = cls["risco"].lower().replace("é", "e")
        bg  = _RISK_COLORS_CSS.get(risco_key, "#FFFFFF")
        fc  = _font_color_css(bg)
        fator   = cls["fator"] or "—"
        limite  = cls["limite_sugerido"] or "Fator × Capacidade de Pagamento"
        lim_rows.append(
            f"<tr>"
            f'<td style="background:{bg};color:{fc};font-weight:bold">{_html.escape(cls["classe"])}</td>'
            f'<td style="background:{bg};color:{fc}">{_html.escape(cls["risco"])}</td>'
            f'<td style="font-weight:bold">{_html.escape(fator)}</td>'
            f'<td style="text-align:left;padding-left:10px">{_html.escape(limite)}</td>'
            f"</tr>"
        )

    limite_section = ""
    if has_limite:
        limite_section = f"""
<div style="margin-top:20px; overflow-x:auto; border:1px solid #D0D7E3; border-radius:8px;
            box-shadow:0 1px 4px rgba(0,0,0,0.08); background:#fff;">
  <table class="cls-tbl" style="max-width:860px;">
    <thead>
      <tr>
        <td colspan="4" style="background:#1F4E79;color:#ffffff;font-family:Calibri,Arial,sans-serif;
            font-size:13px;font-weight:bold;text-align:center;padding:10px;
            border-bottom:1px solid #1A406A;border-radius:8px 8px 0 0;">
          METODOLOGIA DE LIMITE SUGERIDO
        </td>
      </tr>
      <tr>
        <th>Classe</th>
        <th>Risco</th>
        <th>Fator de Limite</th>
        <th>Limite Sugerido</th>
      </tr>
    </thead>
    <tbody>{"".join(lim_rows)}</tbody>
  </table>
</div>"""

    return f"""
{_BASE_CSS}
<style>
  .cls-tbl {{
    border-collapse: collapse;
    font-family: Calibri, Arial, sans-serif;
    font-size: 12px;
    width: 100%;
    max-width: 680px;
    background: #ffffff;
  }}
  .cls-tbl td {{
    border: 1px solid #D0D7E3;
    padding: 7px 14px;
    text-align: center;
    vertical-align: middle;
    color: #000000;
    background: #ffffff;
  }}
  .cls-tbl th {{
    border: 1px solid #1A406A;
    padding: 7px 14px;
    text-align: center;
    background: #1F4E79;
    color: #ffffff;
    font-weight: bold;
  }}
</style>
<div style="overflow-x:auto; border:1px solid #D0D7E3; border-radius:8px;
            box-shadow:0 1px 4px rgba(0,0,0,0.08); background:#fff;">
  <table class="cls-tbl">
    <thead>
      <tr>
        <td colspan="4" style="background:#F0F4FA;color:#1F4E79;font-family:Calibri,Arial,sans-serif;
            font-size:13px;font-weight:bold;text-align:center;padding:10px;
            border-bottom:1px solid #D0D7E3;border-radius:8px 8px 0 0;">
          CLASSIFICACAO DE RISCO &mdash; {_html.escape(model_name)}
        </td>
      </tr>
      <tr>
        <th>Classe</th>
        <th>Risco</th>
        <th>Faixa de Pontuacao</th>
        <th>Validade</th>
      </tr>
    </thead>
    <tbody>{"".join(cls_rows)}</tbody>
  </table>
</div>
{limite_section}
"""


def main() -> None:
    args = sys.argv[1:]
    if not args:
        files = sorted(Path(".").glob("*.json"))
        if not files:
            print("Nenhum .json encontrado.")
            sys.exit(1)
        for f in files:
            try:
                export_json_to_excel(f)
            except Exception as e:
                print(f"Erro em {f.name}: {e}", file=sys.stderr)
        return
    p = Path(args[0])
    if not p.is_file():
        print(f"Arquivo nao encontrado: {p}", file=sys.stderr)
        sys.exit(1)
    export_json_to_excel(p, Path(args[1]) if len(args) > 1 else None)


if __name__ == "__main__":
    main()
